#!/usr/bin/env python3
"""
Options Exchange Fee Schedule Automation
Entry point for CLI and scheduler.

Usage:
    python main.py --preflight                   # Check all endpoints, parse output, no API calls
    python main.py --preflight --exchange edgx   # Preflight a single exchange
    python main.py --run-now                     # All exchanges, full pipeline
    python main.py --mock                        # Same pipeline with mock data (no API key)
    python main.py --mock --mock-jitter          # Mock with simulated rate changes (tests alerts)
    python main.py --exchange edgx               # Single exchange
    python main.py --exchange edgx bzx           # Multiple specific exchanges
    python main.py --schedule                    # Start weekly scheduler (blocking)
    python main.py --report                      # Print cross-exchange fee comparison
    python main.py --excel                       # Export full workbook: exchange sheets + review sheet
    python main.py --excel my_fees.xlsx          # Export to a named file
    python main.py --review                      # Show rows/flags needing human review
    python main.py --footnotes --exchange edgx   # Show footnotes extracted from a schedule
    python main.py --history                     # Print recent run history
"""

from __future__ import annotations

import argparse
import logging
import os
import sys
from pathlib import Path

from dotenv import load_dotenv

load_dotenv()

# Configure logging before importing anything else
LOG_DIR = Path(os.getenv("LOG_DIR", "logs"))
LOG_DIR.mkdir(parents=True, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(LOG_DIR / "fee_automation.log", encoding="utf-8"),
    ],
)
logger = logging.getLogger("main")


def cmd_preflight(exchange_ids: list[str] | None) -> None:
    """Fetch and parse every exchange's fee schedule without calling Claude.

    Checks:
      1. HTTP reachability + status code
      2. Content size (sanity — too small = probably an error page)
      3. Text/CSV/HTML extraction succeeds and produces usable content
      4. Last run status from the DB (so you know if a prior real run worked)
    Prints a colour-coded pass/fail table and a preview of extracted content.
    """
    import time
    import yaml
    from src.fetcher import get_fetcher
    from src.pipeline import _find_manual_file, _manual_fetch_result
    from src.persistence.db import Database

    with open("config/exchanges.yaml") as f:
        cfg = yaml.safe_load(f)

    exchanges = [
        ex for ex in cfg["exchanges"]
        if ex.get("enabled", True)
        and (exchange_ids is None or ex["id"] in exchange_ids)
    ]

    if not exchanges:
        print("No matching enabled exchanges found.")
        return

    db = Database()
    # Most-recent run per exchange (run_history is ordered newest-first)
    last_runs: dict[str, dict] = {}
    for r in reversed(db.get_run_history(limit=200)):
        last_runs[r["exchange_id"]] = r

    PASS  = "\033[92mPASS\033[0m"
    WARN  = "\033[93mWARN\033[0m"
    FAIL  = "\033[91mFAIL\033[0m"
    SKIP  = "\033[90mSKIP\033[0m"

    results = []

    for ex in exchanges:
        xid      = ex["id"]
        xname    = ex["name"]
        url      = ex["fee_url"]
        operator = ex["operator"]

        print(f"\n{'-'*70}")
        print(f"  {xname}  ({xid.upper()})")
        print(f"  URL: {url}")

        t0 = time.time()

        # -- Last run from DB ---------------------------------------------
        last = last_runs.get(xid)
        if last:
            status_icon = PASS if last["status"] == "ok" else FAIL
            print(f"  Last run: {status_icon}  {last['started_at'][:19]}  "
                  f"{last['row_count']} rows  "
                  + (f"ERR: {last['error_message'][:60]}" if last.get("error_message") else ""))
        else:
            print(f"  Last run: {SKIP}  (no prior run in DB)")

        # -- Manual file fallback? ----------------------------------------
        manual = _find_manual_file(xid)
        if manual:
            print(f"  Manual file: {manual}  (will be used instead of HTTP fetch)")

        # -- HTTP fetch ---------------------------------------------------
        fetcher_cls = get_fetcher(operator)
        fetcher = fetcher_cls(ex)
        try:
            if manual:
                fetch_result = _manual_fetch_result(xid, operator, ex, manual)
            else:
                fetch_result = fetcher.fetch()
            elapsed = time.time() - t0
        except Exception as exc:
            print(f"  Fetch: {FAIL}  Exception: {exc}")
            results.append((xid, "FAIL", "fetch exception"))
            continue

        if not fetch_result.ok:
            print(f"  Fetch: {FAIL}  HTTP {fetch_result.http_status}  {fetch_result.error or ''}")
            results.append((xid, "FAIL", f"HTTP {fetch_result.http_status}"))
            continue

        size_kb = len(fetch_result.raw_bytes) / 1024
        print(f"  Fetch: {PASS}  HTTP 200  {size_kb:.1f} KB  {elapsed:.2f}s  "
              f"type={fetch_result.content_type}")

        MIN_BYTES = 500
        if len(fetch_result.raw_bytes) < MIN_BYTES:
            print(f"  {WARN}  Response is suspiciously small ({len(fetch_result.raw_bytes)} bytes) "
                  f"— may be an error page")

        # -- Text extraction ----------------------------------------------
        try:
            text = fetcher.extract_text(fetch_result)
        except Exception as exc:
            print(f"  Parse: {FAIL}  Exception during text extraction: {exc}")
            results.append((xid, "FAIL", "parse exception"))
            continue

        if not text or not text.strip():
            print(f"  Parse: {FAIL}  Extraction returned empty text")
            results.append((xid, "FAIL", "empty text"))
            continue

        char_count = len(text)
        line_count = text.count("\n")
        print(f"  Parse: {PASS}  {char_count:,} chars  {line_count:,} lines")

        # -- Content preview (first 8 non-empty lines) --------------------
        preview_lines = [l for l in text.splitlines() if l.strip()][:8]
        print(f"  Preview:")
        for line in preview_lines:
            print(f"    {line[:100]}")

        # -- Heuristic checks ---------------------------------------------
        warnings = []
        text_lower = text.lower()

        # Check for signs the page returned an error/login wall instead of fee data
        for phrase in ("access denied", "403 forbidden", "sign in", "login required",
                       "page not found", "404", "error occurred",
                       "session expiring", "my nasdaq analyst", "listing center"):
            if phrase in text_lower[:2000]:
                warnings.append(f"response may be a JS login wall (contains '{phrase}')")

        # Check for expected fee-related keywords
        fee_keywords = ("transaction fee", "per contract", "rebate", "maker", "taker",
                        "fee", "rate", "customer", "penny", "code")
        fee_hits = sum(1 for kw in fee_keywords if kw in text_lower)
        if fee_hits < 2:
            warnings.append(
                "very few fee-related keywords found — content may be a JS-rendered page "
                "that requires a browser (Playwright) to load actual fee data"
            )

        if fetch_result.content_type == "csv":
            # CSV-specific: check it has at least 3 columns and multiple rows
            lines = [l for l in text.splitlines() if l.strip()]
            col_counts = [len(l.split(",")) for l in lines[:5]]
            if not all(c >= 2 for c in col_counts):
                warnings.append("CSV has fewer than 2 columns — may be malformed")
            if len(lines) < 5:
                warnings.append(f"CSV has only {len(lines)} rows — suspiciously sparse")
            else:
                print(f"  CSV rows: {len(lines)}")

        for w in warnings:
            print(f"  {WARN}  {w}")

        overall = "WARN" if warnings else "PASS"
        results.append((xid, overall, f"{char_count:,} chars"))

    # -- Summary table ----------------------------------------------------
    print(f"\n{'='*70}")
    print("  PREFLIGHT SUMMARY")
    print(f"{'='*70}")
    passed  = [r for r in results if r[1] == "PASS"]
    warned  = [r for r in results if r[1] == "WARN"]
    failed  = [r for r in results if r[1] == "FAIL"]

    for xid, status, detail in results:
        icon = PASS if status == "PASS" else (WARN if status == "WARN" else FAIL)
        print(f"  {icon}  {xid.upper():<12}  {detail}")

    print(f"\n  {len(passed)} PASS  |  {len(warned)} WARN  |  {len(failed)} FAIL  "
          f"out of {len(results)} checked")

    if failed:
        print(f"\n  Fix FAIL exchanges before running --run-now.")
    elif warned:
        print(f"\n  Review WARN exchanges — they fetched OK but content looks unusual.")
    else:
        print(f"\n  All endpoints healthy. Safe to run --run-now.")


def cmd_run_now(exchange_ids: list[str] | None, mock: bool, mock_jitter: bool) -> None:
    from src.pipeline import run_all
    run_all(exchange_ids=exchange_ids or None, mock=mock, mock_jitter=mock_jitter)


def cmd_schedule() -> None:
    import schedule as sched
    import time
    import yaml

    with open("config/exchanges.yaml") as f:
        cfg = yaml.safe_load(f)

    settings = cfg.get("settings", {})
    run_schedule = settings.get("run_schedule", "weekly")
    run_day = settings.get("run_day", "monday")
    run_time = settings.get("run_time", "06:00")

    from src.pipeline import run_all

    def job():
        logger.info("Scheduler triggered — starting full pipeline run")
        run_all()

    if run_schedule == "daily":
        sched.every().day.at(run_time).do(job)
        logger.info("Scheduler running daily at %s", run_time)
    else:
        day_fn = getattr(sched.every(), run_day)
        day_fn.at(run_time).do(job)
        logger.info("Scheduler running every %s at %s", run_day, run_time)

    logger.info("Press Ctrl+C to stop")
    while True:
        sched.run_pending()
        time.sleep(60)


def cmd_report(trade_type_filter: str | None = None) -> None:
    import json as _json
    from src.persistence.db import Database
    from tabulate import tabulate

    db = Database()
    rows = db.get_all_latest_rows()

    if not rows:
        print("No data in database. Run --run-now or --mock first.")
        return

    _sec_order   = {"OPT": 0, "MLEG": 1}
    _acct_order  = {"CUST": 0, "PCUST": 1}
    _trade_order = {"Electronic": 0, "PI": 1, "Solicitation": 2}
    _class_order = {"Penny": 0, "Non-Penny": 1}

    sort_key = lambda x: (
        x.get("exchange_id") or "",
        _sec_order.get(x.get("sec_type") or "", 9),
        _acct_order.get(x.get("account_type") or "", 9),
        _trade_order.get(x.get("trade_type") or "", 9),
        x.get("liq_code") or "",
        _class_order.get(x.get("ticker_class") or "", 9),
    )

    RATE_FIELDS = ["make_rate", "take_rate", "auction_init_rate", "auction_resp_rate", "breakup_rate"]
    headers = ["Exchange", "LiqCode", "SecType", "AcctType", "TradeType", "Class",
               "Make", "Take", "AuctInit", "AuctResp", "Breakup", "Cf", "FnRefs"]

    display_rows = []

    for r in sorted(rows, key=sort_key):
        if trade_type_filter and r.get("trade_type") != trade_type_filter:
            continue
        fn_refs = _json.loads(r.get("footnote_refs") or "[]")
        conf = r.get("confidence", "high")
        conf_mark = "" if conf == "high" else ("?" if conf == "medium" else "!")
        fn_mark = ",".join(fn_refs) if fn_refs else ""

        meta = [
            (r.get("exchange_id") or "").upper(),
            r.get("liq_code") or "",
            r.get("sec_type") or "",
            r.get("account_type") or "",
            r.get("trade_type") or "",
            r.get("ticker_class") or "",
        ]
        display_rows.append(meta + [_fmt(r.get(f)) for f in RATE_FIELDS] + [conf_mark, fn_mark])

    print(tabulate(display_rows, headers=headers, tablefmt="outline"))
    exchange_count = len(set(r["exchange_id"] for r in rows))
    filter_note = f" ({trade_type_filter} only)" if trade_type_filter else ""
    print(f"\n{len(display_rows)} rows shown{filter_note} | {len(rows)} total rows across {exchange_count} exchange(s)")
    print("Cf: blank=high confidence, ?=medium, !=low  |  FnRefs: applicable footnote IDs")


def cmd_excel(path: str = "fee_schedule.xlsx") -> None:
    """Export a single workbook: one sheet per exchange (fees + FnNotes) + a Review sheet."""
    import json as _json
    from src.persistence.db import Database
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from collections import defaultdict

    db = Database()
    all_rows = db.get_all_latest_rows()

    if not all_rows:
        print("No data in database. Run --run-now or --mock first.")
        return

    # ---- shared style constants ----
    RATE_COLS    = {"Make", "Take", "AuctInit", "AuctResp", "Breakup"}
    RATE_FMT     = '#,##0.00;[Red]-#,##0.00'
    HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
    FEE_HDR      = PatternFill("solid", fgColor="1F4E79")   # dark blue — fee sheets
    REV_HDR      = PatternFill("solid", fgColor="7B2D2D")   # dark red  — review sheet
    ALT_FILL     = PatternFill("solid", fgColor="D6E4F0")
    LOW_FILL     = PatternFill("solid", fgColor="FCE4D6")
    MED_FILL     = PatternFill("solid", fgColor="FFF2CC")
    ALT_MED_FILL = PatternFill("solid", fgColor="FFEB9C")

    def _autosize(ws, headers, sheet_rows, cap=60):
        for col_idx, h in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            max_len = max(
                (len(str(sheet_rows[i][col_idx - 1] or "")) for i in range(len(sheet_rows))),
                default=0,
            )
            ws.column_dimensions[col_letter].width = min(max(max_len, len(h)) + 2, cap)

    def _write_header(ws, headers, fill):
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = HEADER_FONT
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"

    # ---- sort order ----
    _sec   = {"OPT": 0, "MLEG": 1}
    _acct  = {"CUST": 0, "PCUST": 1}
    _trade = {"Electronic": 0, "PI": 1, "Solicitation": 2}
    _cls   = {"Penny": 0, "Non-Penny": 1}
    sort_key = lambda x: (
        (x.get("exchange_id") or "").lower(),
        _sec.get(x.get("sec_type") or "", 9),
        _acct.get(x.get("account_type") or "", 9),
        _trade.get(x.get("trade_type") or "", 9),
        x.get("liq_code") or "",
        _cls.get(x.get("ticker_class") or "", 9),
    )

    RATE_FIELDS   = ["make_rate", "take_rate", "auction_init_rate", "auction_resp_rate", "breakup_rate"]
    FEE_HEADERS   = ["LiqCode", "SecType", "AcctType", "TradeType", "Class",
                     "Make", "Take", "AuctInit", "AuctResp", "Breakup", "Cf", "FnRefs", "FnNotes"]

    fn_lookup = db.get_all_latest_footnotes()

    # ---- build review rows first so we can map keys → Review sheet row numbers ----
    review_db_rows = db.get_review_needed()
    REV_HEADERS  = ["Exchange", "LiqCode", "Conf", "AcctType", "Class",
                    "SecType", "TradeType", "Make", "Take",
                    "AuctInit", "AuctResp", "Breakup", "Source Citation", "Reason"]
    REV_CONF_IDX = REV_HEADERS.index("Conf")
    review_rows  = []
    # key: (exchange_upper, liq_code, sec_type, acct_type, trade_type, class) → Excel row number on Review sheet
    review_row_map: dict[tuple, int] = {}
    for r in review_db_rows:
        fn_refs  = _json.loads(r.get("footnote_refs") or "[]")
        citation = " > ".join(filter(None, [
            r.get("source_page"), r.get("source_section"),
            (f"fn. {', '.join(fn_refs)}" if fn_refs else None),
        ])) or "unknown"
        review_rows.append([
            (r.get("exchange_id") or "").upper(),
            r.get("liq_code") or "",
            (r.get("confidence") or "").upper(),
            r.get("account_type") or "",
            r.get("ticker_class") or "",
            r.get("sec_type") or "",
            r.get("trade_type") or "",
        ] + [
            float(r[f]) if r.get(f) is not None else None for f in RATE_FIELDS
        ] + [citation, r.get("confidence_reason") or ""])
        key = (
            (r.get("exchange_id") or "").upper(),
            r.get("liq_code") or "",
            r.get("sec_type") or "",
            r.get("account_type") or "",
            r.get("trade_type") or "",
            r.get("ticker_class") or "",
        )
        if key not in review_row_map:
            review_row_map[key] = len(review_rows) + 1  # +1 for header row → Excel row number

    # ---- build per-exchange fee rows ----
    # each entry: (xid_upper, fee_row_list)
    by_exchange: dict[str, list] = defaultdict(list)
    all_fee_rows_with_xid: list[tuple[str, list]] = []  # for All sheet
    for r in sorted(all_rows, key=sort_key):
        xid      = (r.get("exchange_id") or "").lower()
        xid_up   = xid.upper()
        fn_refs  = _json.loads(r.get("footnote_refs") or "[]")
        conf     = r.get("confidence", "high")
        fn_notes = "  |  ".join(
            f"[{ref}] {fn_lookup[(xid, ref)]}"
            for ref in fn_refs if (xid, ref) in fn_lookup
        )
        cf_mark = "" if conf == "high" else ("?" if conf == "medium" else "!")
        row = [
            r.get("liq_code") or "",
            r.get("sec_type") or "",
            r.get("account_type") or "",
            r.get("trade_type") or "",
            r.get("ticker_class") or "",
        ] + [
            float(r[f]) if r.get(f) is not None else None for f in RATE_FIELDS
        ] + [cf_mark, ",".join(fn_refs), fn_notes]
        by_exchange[xid_up].append(row)
        all_fee_rows_with_xid.append((xid_up, row))

    CF_COL_IDX = FEE_HEADERS.index("Cf") + 1          # 1-based column number of Cf in exchange sheets
    ALL_CF_COL_IDX = CF_COL_IDX + 1                    # All sheet has Exchange prepended → shift by 1
    LINK_FONT    = Font(color="0563C1", underline="single", bold=False, size=10)
    ALL_HEADERS  = ["Exchange"] + FEE_HEADERS

    def _write_fee_sheet(ws, headers, sheet_rows, cf_col, xid_for_lookup):
        """Write a fee sheet. cf_col is the 1-based Cf column index."""
        _write_header(ws, headers, FEE_HDR)
        for row_idx, row in enumerate(sheet_rows, 2):
            # For All sheet row is (xid, fee_row); for exchange sheet it's just fee_row
            if xid_for_lookup is None:
                actual_xid, data_row = row[0], row[1]
            else:
                actual_xid, data_row = xid_for_lookup, row
            fill = ALT_FILL if row_idx % 2 == 0 else None
            display_row = ([actual_xid] + data_row) if xid_for_lookup is None else data_row
            for col_idx, val in enumerate(display_row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                if fill:
                    cell.fill = fill
                if headers[col_idx - 1] in RATE_COLS and val is not None:
                    cell.number_format = RATE_FMT
            # Hyperlink on Cf cell if non-empty and row is in review map
            cf_val = display_row[cf_col - 1]
            if cf_val:
                liq_code  = data_row[FEE_HEADERS.index("LiqCode")]
                sec_type  = data_row[FEE_HEADERS.index("SecType")]
                acct_type = data_row[FEE_HEADERS.index("AcctType")]
                trade     = data_row[FEE_HEADERS.index("TradeType")]
                cls_      = data_row[FEE_HEADERS.index("Class")]
                key = (actual_xid, liq_code, sec_type, acct_type, trade, cls_)
                if key in review_row_map:
                    rev_row = review_row_map[key]
                    cf_cell = ws.cell(row=row_idx, column=cf_col)
                    cf_cell.hyperlink = f"#Review!A{rev_row}"
                    cf_cell.font = LINK_FONT
        _autosize(ws, headers, [
            ([r[0]] + list(r[1])) if xid_for_lookup is None else r
            for r in sheet_rows
        ])

    # ---- assemble workbook ----
    wb = Workbook()
    wb.remove(wb.active)

    # Per-exchange fee sheets
    for xid in sorted(by_exchange):
        ws = wb.create_sheet(title=xid)
        _write_fee_sheet(ws, FEE_HEADERS, by_exchange[xid], CF_COL_IDX, xid)

    # Review sheet
    if review_rows:
        ws_rev = wb.create_sheet(title="Review")
        _write_header(ws_rev, REV_HEADERS, REV_HDR)
        for row_idx, row in enumerate(review_rows, 2):
            conf = str(row[REV_CONF_IDX]).upper()
            fill = LOW_FILL if conf == "LOW" else (MED_FILL if row_idx % 2 == 0 else ALT_MED_FILL)
            for col_idx, val in enumerate(row, 1):
                cell = ws_rev.cell(row=row_idx, column=col_idx, value=val)
                cell.fill = fill
                if REV_HEADERS[col_idx - 1] in RATE_COLS and val is not None:
                    cell.number_format = RATE_FMT
        _autosize(ws_rev, REV_HEADERS, review_rows)

    # All sheet at the front
    ws_all = wb.create_sheet(title="All", index=0)
    _write_fee_sheet(ws_all, ALL_HEADERS, all_fee_rows_with_xid, ALL_CF_COL_IDX, None)

    wb.save(path)
    exchange_count = len(by_exchange)
    review_count   = len(review_rows)
    print(f"Workbook written to: {path}")
    print(f"  All sheet + {exchange_count} exchange sheet(s) | {review_count} review row(s)")


def cmd_review() -> None:
    """Print all rows and flags that need human review, with source citations."""
    from src.persistence.db import Database
    from tabulate import tabulate
    import json

    def _t(s: str | None, n: int) -> str:
        s = s or ""
        return s if len(s) <= n else s[: n - 1] + "…"

    db = Database()

    flags = db.get_flags()
    if flags:
        print(f"\n{'='*60}")
        print("EXTRACTION FLAGS (AI-identified issues)")
        print('='*60)
        for f in flags:
            print(f"  [{f['severity'].upper()}] {f['exchange_id'].upper()} | {f['location']}")
            print(f"  {f['issue']}")
            print()

    rows = db.get_review_needed()
    if not rows and not flags:
        print("Nothing needs review — all rows extracted at high confidence.")
        return

    if rows:
        print(f"\n{'='*60}")
        print("LOW / MEDIUM CONFIDENCE ROWS")
        print('='*60)
        table = []
        for r in rows:
            fn_refs = json.loads(r.get("footnote_refs") or "[]")
            citation = " > ".join(filter(None, [
                r.get("source_page"), r.get("source_section"),
                (f"fn. {', '.join(fn_refs)}" if fn_refs else None),
            ])) or "unknown"
            table.append([
                r["exchange_id"].upper(),
                r.get("liq_code") or "",
                r.get("confidence", "").upper(),
                r.get("account_type", ""),
                r.get("ticker_class", "") or "",
                r.get("sec_type", ""),
                r.get("trade_type", ""),
                _t(citation, 38),
                _t(r.get("confidence_reason") or "", 45),
            ])
        print(tabulate(table,
                       headers=["Exchange", "LiqCode", "Conf", "AcctType", "Class",
                                 "SecType", "TradeType", "Source Citation", "Reason"],
                       tablefmt="outline"))

    print(f"\n{len(rows)} row(s) need review | {len(flags)} flag(s) total")


def cmd_footnotes(exchange_id: str) -> None:
    """Print all footnotes extracted from the latest run for an exchange."""
    import yaml
    from src.persistence.db import Database

    with open("config/exchanges.yaml") as f:
        cfg = yaml.safe_load(f)
    ex = next((e for e in cfg["exchanges"] if e["id"] == exchange_id), None)

    db = Database()
    footnotes = db.get_footnotes(exchange_id)

    # Explain the CBOE CSV footnote limitation
    if not footnotes and ex and ex.get("operator") == "cboe":
        print(
            f"\nNo footnotes stored for {exchange_id.upper()}.\n\n"
            f"CBOE exchanges are fetched via a CSV export endpoint that strips all footnotes\n"
            f"from the underlying fee schedule. The CBOE fee schedule website at\n"
            f"  cboe.com/us/options/membership/fee_schedule/{exchange_id}/\n"
            f"contains footnotes qualifying volume thresholds, program eligibility, and rate\n"
            f"caps that are NOT reflected in the CSV. This is why all CBOE rows are marked\n"
            f"medium confidence — the CSV alone is insufficient for full footnote coverage.\n\n"
            f"To review CBOE footnotes manually, visit the fee schedule page above."
        )
        return

    if not footnotes:
        print(f"No footnotes found for {exchange_id.upper()}. Run --run-now first.")
        return

    print(f"\nFootnotes extracted from {exchange_id.upper()} fee schedule "
          f"({len(footnotes)} total):\n")
    for fn in footnotes:
        print(f"  [{fn['ref']}]  Location: {fn['location']}")
        print(f"       {fn['text']}")
        print()


def cmd_preview_alerts(preview_dir: "Path | None" = None) -> None:
    """Write all alert card types to data/alert-preview/ using synthetic data.

    Teams JSON files can be pasted into https://adaptivecards.io/designer
    to see the exact card rendering before configuring a live webhook.
    """
    from pathlib import Path as _Path
    from src.alerts.teams import TeamsAlerter
    from src.alerts.email import EmailAlerter
    from src.diff.engine import DiffReport, RowChange, RateChange

    out_dir = _Path(preview_dir) if preview_dir is not None else _Path("data/alert-preview")
    out_dir.mkdir(parents=True, exist_ok=True)

    teams = TeamsAlerter(webhook_url="", dry_run=True, preview_dir=out_dir)
    email = EmailAlerter(dry_run=True, preview_dir=out_dir)

    # --- synthetic data ---
    report_changed = DiffReport(
        exchange_id="edgx",
        has_changes=True,
        modified=[
            RowChange(
                key={"exchange_id": "edgx", "ticker_class": "Penny", "sec_type": "OPT",
                     "account_type": "CUST", "trade_type": "Electronic", "liq_code": "CA"},
                change_type="modified",
                rate_changes=[
                    RateChange(field="make_rate", field_label="Make Rate",
                               old_value=-0.48, new_value=-0.50),
                    RateChange(field="take_rate", field_label="Take Rate",
                               old_value=-0.48, new_value=-0.46),
                ],
            ),
            RowChange(
                key={"exchange_id": "edgx", "ticker_class": "Non-Penny", "sec_type": "OPT",
                     "account_type": "PCUST", "trade_type": "PI", "liq_code": "PA"},
                change_type="modified",
                rate_changes=[
                    RateChange(field="auction_init_rate", field_label="Auction Init Rate",
                               old_value=-0.20, new_value=-0.22),
                ],
            ),
        ],
        added=[
            RowChange(
                key={"exchange_id": "edgx", "ticker_class": "Non-Penny", "sec_type": "OPT",
                     "account_type": "CUST", "trade_type": "Electronic", "liq_code": "NA"},
                change_type="added",
            ),
        ],
        removed=[
            RowChange(
                key={"exchange_id": "edgx", "ticker_class": "Penny", "sec_type": "MLEG",
                     "account_type": "CUST", "trade_type": "Solicitation", "liq_code": "SL"},
                change_type="removed",
            ),
        ],
    )
    report_unchanged = DiffReport(exchange_id="bzx", has_changes=False)
    all_reports = [report_changed, report_unchanged]

    files_written = []

    # Teams: per-exchange diff (with changes)
    teams.send_diff_report(report_changed)
    files_written.append(out_dir / "teams_diff_edgx.json")

    # Teams: per-exchange no change
    teams.send_diff_report(report_unchanged)
    files_written.append(out_dir / "teams_no_change_bzx.json")

    # Teams: error card
    teams.send_error("c2", "HTTP 403 — CDN blocked request after 3 retries")
    files_written.append(out_dir / "teams_error_c2.json")

    # Teams: full run summary (mix of changed, unchanged, error)
    teams.send_run_summary(
        all_reports,
        errors=[("c2", "HTTP 403 — CDN blocked")],
        cross_exchange_insight=(
            "EDGX increased make fees by $0.02/contract on Penny options, "
            "consistent with a broader industry trend toward tighter rebates this quarter."
        ),
    )
    files_written.append(out_dir / "teams_run_summary.json")

    # Teams: review-needed card
    teams.send_review_needed(
        "edgx",
        "CA (Penny/OPT/CUST/Electronic): make_rate extracted as null — "
        "fee code description ambiguous\n"
        "PA (Non-Penny/OPT/PCUST/PI): source_section missing",
    )
    files_written.append(out_dir / "teams_review_edgx.json")

    # Email: per-exchange diff
    email.send_diff_report(report_changed)
    files_written.append(out_dir / "email_diff_edgx.txt")

    # Email: run summary
    email.send_run_summary(all_reports, errors=[("c2", "HTTP 403")])
    files_written.append(out_dir / "email_run_summary.txt")

    print(f"\nAlert previews written to: {out_dir.resolve()}")
    print(f"\n  Teams cards ({len([f for f in files_written if f.suffix == '.json'])} files):")
    for f in files_written:
        if f.suffix == ".json":
            print(f"    {f.name}")
    print(f"\n  Email previews ({len([f for f in files_written if f.suffix == '.txt'])} files):")
    for f in files_written:
        if f.suffix == ".txt":
            print(f"    {f.name}")
    print(
        "\n  To preview Teams cards: open https://adaptivecards.io/designer"
        "\n  paste the JSON from any teams_*.json file into the 'Card Payload Editor' panel."
    )


def cmd_history(limit: int = 20) -> None:
    from src.persistence.db import Database
    from tabulate import tabulate

    db = Database()
    runs = db.get_run_history(limit=limit)

    if not runs:
        print("No run history found.")
        return

    table = [
        [
            r["run_id"],
            r["exchange_id"].upper(),
            r["started_at"][:19],
            r["status"],
            r["row_count"],
            r.get("error_message", "")[:60] if r.get("error_message") else "",
        ]
        for r in runs
    ]
    headers = ["RunID", "Exchange", "Started", "Status", "Rows", "Error"]
    print(tabulate(table, headers=headers, tablefmt="outline"))


def _fmt(val) -> str:
    if val is None:
        return "-"
    v = float(val)
    sign = "+" if v >= 0 else ""
    return f"{sign}${v:.2f}"


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Options Exchange Fee Schedule Automation",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument(
        "--preflight",
        action="store_true",
        help="Fetch and parse every endpoint without calling Claude — verify before spending tokens",
    )
    group.add_argument(
        "--run-now",
        action="store_true",
        help="Run the full pipeline immediately (requires ANTHROPIC_API_KEY)",
    )
    group.add_argument(
        "--mock",
        action="store_true",
        help="Run with synthetic mock data — no API key required, tests the full pipeline",
    )
    group.add_argument(
        "--schedule",
        action="store_true",
        help="Start the scheduler (blocking)",
    )
    group.add_argument(
        "--report",
        action="store_true",
        help="Print cross-exchange fee comparison from stored data",
    )
    group.add_argument(
        "--history",
        action="store_true",
        help="Print recent run history",
    )
    group.add_argument(
        "--review",
        action="store_true",
        help="Show rows and flags that need human review, with source citations",
    )
    group.add_argument(
        "--footnotes",
        action="store_true",
        help="Show all footnotes extracted from the latest run (requires --exchange)",
    )
    group.add_argument(
        "--excel",
        nargs="?",
        const="fee_schedule.xlsx",
        metavar="FILE",
        help="Export all exchange fee data + review sheet to a single Excel workbook "
             "(default: fee_schedule.xlsx)",
    )
    group.add_argument(
        "--preview-alerts",
        action="store_true",
        help="Write all alert card types to data/alert-preview/ using synthetic data — "
             "paste Teams JSON into adaptivecards.io/designer to preview cards",
    )

    parser.add_argument(
        "--mock-jitter",
        action="store_true",
        help="With --mock: randomly perturb some rates to simulate changes and trigger alerts",
    )
    parser.add_argument(
        "--exchange",
        nargs="+",
        metavar="ID",
        help="Limit to specific exchange ID(s), e.g. --exchange edgx bzx",
    )
    parser.add_argument(
        "--filter-trade-type",
        choices=["Electronic", "PI", "Solicitation"],
        metavar="TYPE",
        help="Filter --report to one trade type: Electronic, PI, or Solicitation",
    )
    parser.add_argument(
        "--debug",
        action="store_true",
        help="Enable debug logging",
    )

    args = parser.parse_args()

    if args.debug:
        logging.getLogger().setLevel(logging.DEBUG)

    if args.mock_jitter and not args.mock:
        parser.error("--mock-jitter requires --mock")

    if args.preflight:
        cmd_preflight(args.exchange)
    elif args.run_now:
        cmd_run_now(args.exchange, mock=False, mock_jitter=False)
    elif args.mock:
        cmd_run_now(args.exchange, mock=True, mock_jitter=args.mock_jitter)
    elif args.schedule:
        if args.exchange:
            parser.error("--exchange cannot be used with --schedule")
        cmd_schedule()
    elif args.report:
        cmd_report(trade_type_filter=getattr(args, "filter_trade_type", None))
    elif args.history:
        cmd_history()
    elif args.review:
        cmd_review()
    elif args.excel is not None:
        cmd_excel(args.excel)
    elif args.footnotes:
        if not args.exchange or len(args.exchange) != 1:
            parser.error("--footnotes requires exactly one --exchange ID")
        cmd_footnotes(args.exchange[0])
    elif args.preview_alerts:
        cmd_preview_alerts()


if __name__ == "__main__":
    main()
